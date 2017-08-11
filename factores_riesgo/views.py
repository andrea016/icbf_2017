#!/usr/bin/env python
# -*- coding: utf-8 -*-

from django.http import HttpResponseRedirect, HttpResponse,Http404
from django.template import RequestContext
from django.db import transaction
from django.shortcuts import render, render_to_response
from django.contrib import auth, messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.csrf import csrf_exempt
import xhtml2pdf.pisa as pisa
from StringIO import StringIO
from django.template.loader import render_to_string
from django.db.models import Q
from remax.settings import URL
from contactos.models import *
from calendario.models import *
from sucursales.models import *
from referidos.models import Referido
from agentes.models import Agente


########## LISTA DE CONTACTOS ###########

from django.contrib import messages
@login_required(login_url="login:login")
def contactos(request):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        usuarios = User.objects.all()
        contactos = Contacto.objects.filter(agen_contacto=request.user.id).order_by('nombres')
        return render(request,'contactos/contactos_list.html',{'contactos': contactos })

########## TEMPLATE  CREAR  CONTACTO ###########

@login_required(login_url="login:login")
def crearContacto(request):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        paises = Paises.objects.all()
        t_contactos = Tipo_contacto.objects.filter(agente=request.user.id).order_by('tipo')
        contactos = Contacto.objects.filter(agen_contacto=request.user.id).order_by('nombres')
        referidos = Referido.objects.all()
        profesiones = Profesiones.objects.all()
        t_residencias = Tipo_residencia.objects.all()
        t_direcciones = Tipo_direccion.objects.all()
        t_viviendas = Tipo_vivienda.objects.all()
        return render(request,'contactos/nuevo_contacto.html',{'paises': paises ,'t_contactos': t_contactos,'contactos': contactos,'profesiones': profesiones,'t_residencias': t_residencias,'t_direcciones': t_direcciones,'t_viviendas': t_viviendas, 'referidos': referidos})

########## FUNCION GUARDAR CONTACTO ###########

@transaction.atomic
@csrf_exempt
@login_required(login_url="login:login")
def guardarContacto(request):
    if request.method == 'POST':
        if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
            return HttpResponseRedirect('/')
        else:
            contacto = Contacto()
            n = Notas()
            contacto.nombres = request.POST['nombres'].encode('ascii', 'ignore')
            contacto.apellidos = request.POST['apellidos'].encode('ascii', 'ignore')
            contacto.correo = request.POST['correo']
            contacto.genero = request.POST['genero']
            contacto.dni = request.POST['documento']
            contacto.agen_contacto= int(request.user.id)

            if request.POST['date'] != "":
                contacto.nacimiento = request.POST['date']

            if request.POST['bandera_foto'] == "CAMBIO":
                contacto.foto = request.FILES['archivo']
            else:
                contacto.foto = "contactos/no_photo.png"

            contacto.agen_contacto = request.user.id

            if request.POST['referido'] != "":
                contacto.referido_id = int(request.POST['referido'])
            if request.POST['t_contacto'] != "":
                contacto.t_contacto_id = int(request.POST['t_contacto'])

            contacto.palabras = request.POST['palabras_clave'].encode('ascii', 'ignore')
            contacto.celular = request.POST['celular']
            contacto.t_casa = request.POST['t_casa']

            if request.POST['profesion_id'] != "":
                contacto.profesion_id = int(request.POST['profesion_id'])

            contacto.empresa = request.POST['empresa'].encode('ascii', 'ignore')
            contacto.ruc_empresa = request.POST['ruc'].encode('ascii', 'ignore')
            contacto.t_oficina = request.POST['t_oficina']
            contacto.ext = request.POST['ext']
            contacto.fax = request.POST['fax']
            contacto.sitio_web = request.POST['sitio_web']

            if request.POST['pais'] != "":
                contacto.pais_id =  int(request.POST['pais'])
            if request.POST['departamento'] != "":
                contacto.departamento_id = int(request.POST['departamento'])
            if request.POST['provincia'] != "":
                contacto.provincia_id = int(request.POST['provincia'])
            if request.POST['distrito'] != "":
                contacto.distrito_id= int(request.POST['distrito'])
            if request.POST['ciudad'] != "":
                contacto.ciudad_id = int(request.POST['ciudad'])
            if request.POST['t_residencia_id'] != "":
                contacto.t_residencia_id = int(request.POST['t_residencia_id'])


            contacto.residencia = request.POST['residencia'].encode('ascii', 'ignore')

            if request.POST['t_direccion_id'] != "":
                contacto.t_direccion_id = int(request.POST['t_direccion_id'])

            contacto.direccion = request.POST['direccion'].encode('ascii', 'ignore')
            contacto.num_dir = request.POST['num_dir']
            contacto.manzana = request.POST['mzn_dir']
            contacto.num_lote = request.POST['num_lote']

            if request.POST['t_vivienda_id'] != "":
                contacto.t_vivienda_id = int(request.POST['t_vivienda_id'])

            contacto.dep_num = request.POST['nro_vivienda']
            contacto.cod_postal = request.POST['cod_postal']
            contacto.saludo = request.POST['saludo_id']
            contacto.carta = request.POST['carta_id']

            if request.POST['conyugue_id'] != "":
                contacto.conyugue = request.POST['conyugue_id']
                contacto.save()

                # ASIGNO QUE EL OTRO USUARIO QUE ES CONYUGE DE ESTE
                c = Contacto.objects.get(id = contacto.conyugue)
                c.conyugue = contacto.id
                c.save()

            contacto.save()

            n.agente = request.user.id
            n.usuario = contacto.id
            mensaje = ("Creacion del Contacto")
            n.nota = mensaje.encode('ascii', 'ignore')
            n.save()

            c = Contacto.objects.get(id=contacto.id)
            c.nota_id = n.id
            c.save()

            messages.success(request, 'Creado')
            return HttpResponseRedirect('/contactos')
    else:
        return HttpResponseRedirect("/")

########## TEMPLATE EDITAR CONTACTO ###########

@login_required(login_url="login:login")
def editarContacto(request,id = None):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        admin = User.objects.get(groups__name='administradores')
        paises = Paises.objects.all()
        t_contactos = Tipo_contacto.objects.filter(agente=request.user.id).order_by('tipo')
        contacto = Contacto.objects.get(id = id)
        referidos = Referido.objects.all()
        conyugues = Contacto.objects.exclude( id = id)
        profesiones = Profesiones.objects.all()
        t_residencias = Tipo_residencia.objects.all()
        t_direcciones = Tipo_direccion.objects.all()
        t_viviendas = Tipo_vivienda.objects.all()
        url = URL
        return render(request,'contactos/editar_contacto.html',{'paises': paises ,'t_contactos': t_contactos,'contacto': contacto,'conyugues': conyugues,'profesiones': profesiones,'t_residencias': t_residencias,'t_direcciones': t_direcciones,'t_viviendas': t_viviendas, 'url': url , 'referidos' : referidos})


############## FUNCION ACTUALIZAR CONTACTO #######################

@login_required(login_url="login:login")
def actualizarContacto(request):
    if request.method == 'POST':
        if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
                return HttpResponseRedirect('/')
        else:
            id = request.POST['contacto_id']
            contacto = Contacto.objects.get(id = id)
            contacto.nombres = request.POST['nombres'].encode('ascii', 'ignore')
            contacto.apellidos = request.POST['apellidos'].encode('ascii', 'ignore')
            contacto.correo = request.POST['correo']
            contacto.genero = request.POST['genero']
            contacto.dni = request.POST['documento']

            if request.POST['date'] != "":
                contacto.nacimiento = request.POST['date']

            f_anterior = str(contacto.foto)
            if request.POST['bandera_foto'] == "CAMBIO":
                contacto.foto  = request.FILES['archivo']
                if  f_anterior != ("contactos/no_photo.png"):
                    os.remove("media/"+f_anterior)
                else:
                    pass
            else:
                contacto.foto = f_anterior

            if request.POST['referido'] != "":
                contacto.referido_id = int(request.POST['referido'])

            if request.POST['t_contacto'] != "":
                contacto.t_contacto_id = int(request.POST['t_contacto'])

            if request.POST['conyugue_id'] != "":
                contacto.conyugue = request.POST['conyugue_id']

            contacto.palabras = request.POST['palabras_clave'].encode('ascii', 'ignore')
            contacto.celular = request.POST['celular']
            contacto.t_casa = request.POST['t_casa']

            if request.POST['profesion_id'] != "":
                contacto.profesion_id = int(request.POST['profesion_id'])

            contacto.empresa = request.POST['empresa'].encode('ascii', 'ignore')
            contacto.ruc_empresa = request.POST['ruc'].encode('ascii', 'ignore')
            contacto.t_oficina = request.POST['t_oficina']
            contacto.ext = request.POST['ext']
            contacto.fax = request.POST['fax']
            contacto.sitio_web = request.POST['sitio_web']

            if request.POST['pais'] != "":
                contacto.pais_id =  int(request.POST['pais'])

            if request.POST['departamento'] != "":
                contacto.departamento_id = int(request.POST['departamento'])

            if request.POST['provincia'] != "":
                contacto.provincia_id = int(request.POST['provincia'])

            if request.POST['distrito'] != "":
                contacto.distrito_id= int(request.POST['distrito'])

            if request.POST['ciudad'] != "":
                contacto.ciudad_id = int(request.POST['ciudad'])

            if request.POST['t_residencia_id'] != "":
                contacto.t_residencia_id = int(request.POST['t_residencia_id'])

            contacto.residencia = request.POST['residencia'].encode('ascii', 'ignore')

            if request.POST['t_direccion_id'] != "":
                contacto.t_direccion_id = int(request.POST['t_direccion_id'])

            contacto.direccion = request.POST['direccion'].encode('ascii', 'ignore')
            contacto.num_dir = request.POST['num_dir']
            contacto.manzana = request.POST['mzn_dir']
            contacto.num_lote = request.POST['num_lote']

            if request.POST['t_vivienda_id'] != "":
                contacto.t_vivienda_id = int(request.POST['t_vivienda_id'])

            contacto.dep_num = request.POST['nro_vivienda']
            contacto.cod_postal = request.POST['cod_postal']
            contacto.saludo = request.POST['saludo_id']
            contacto.carta = request.POST['carta_id']


            if request.POST['conyugue_id'] != "":
                contacto.conyugue = request.POST['conyugue_id']
                contacto.save()

                # ASIGNO QUE EL OTRO USUARIO QUE ES CONYUGE DE ESTE
                c = Contacto.objects.get(id = contacto.conyugue)
                c.conyugue = contacto.id
                c.save()
            else:
                contacto.conyugue = ""


            contacto.save()
            messages.success(request, 'Actualizado')
            return HttpResponseRedirect('/contactos')
    else:
        return HttpResponseRedirect('/')


########### TEMPLATE  REPORTE DE CONTACTO ###############

@login_required(login_url="login:login")
def reporteContacto(request,id = None):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        contacto = Contacto.objects.get(id = id)
        tareas = Eventos.objects.filter(tipo="TAREAS").filter(contactos=id).order_by('f_inicio')
        pendientes = Eventos.objects.filter(tipo="PENDIENTES").filter(contactos=id).order_by('f_inicio')
        llamadas = Eventos.objects.filter(tipo="LLAMADAS").filter(contactos=id).order_by('f_inicio')
        citas = Eventos.objects.filter(tipo="CITAS").filter(contactos=id).order_by('f_inicio')
        propiedades = Eventos.objects.filter(Q(tipo="MOSTRADAS") | Q(tipo="PROPIEDADES")).filter(contactos=id).order_by('f_inicio')
        notas = Notas.objects.filter(usuario=id).order_by('fecha')
        url = URL
        return render(request,'contactos/reporte_contacto.html',{'contacto': contacto, 'url': url, 'tareas': tareas, 'pendientes': pendientes,'llamadas': llamadas, 'citas': citas, 'propiedades': propiedades, 'notas': notas })


############## FUNCION REPORTE DE CONTACTO #######################

@login_required(login_url="login:login")
def ContactoPDF(request, id=None):

    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        url = URL
        logo = "remax-peru.png"
        contacto = Contacto.objects.get(id = id)
        agente =  Agente.objects.get(id = request.user.id)
        sucursal = Sucursales.objects.get(id=agente.sucursal_id)
        tareas = Eventos.objects.filter(tipo="TAREAS").filter(contactos=id).order_by('f_inicio')
        pendientes = Eventos.objects.filter(tipo="PENDIENTES").filter(contactos=id).order_by('f_inicio')
        llamadas = Eventos.objects.filter(tipo="LLAMADAS").filter(contactos=id).order_by('f_inicio')
        citas = Eventos.objects.filter(tipo="CITAS").filter(contactos=id).order_by('f_inicio')
        propiedades = Eventos.objects.filter(tipo="PROPIEDADES").filter(contactos=id).order_by('f_inicio')
        mostradas = Eventos.objects.filter(tipo="MOSTRADAS").filter(contactos=id).order_by('f_inicio')
        notas = Notas.objects.filter(usuario=id).order_by('fecha')
        result = StringIO()
        html= render_to_string("contactos/contactos_pdf.html",{"url": url, "logo": logo, "contacto": contacto, "tareas": tareas, "pendientes": pendientes, "llamadas": llamadas,"citas": citas, "propiedades": propiedades ,"mostradas": mostradas, "notas": notas , "sucursal": sucursal})
        pdf = pisa.pisaDocument(html,result)

        response = HttpResponse(result.getvalue(),content_type='application/pdf')
        name = contacto.nombres+" "+contacto.apellidos+".pdf"
        response['Content-Disposition'] = "attachment; filename='%s'" % name
        return response


############## FUNCION ELIMINAR CONTACTO #######################

@csrf_exempt
@transaction.atomic
@login_required(login_url="login:login")
def eliminarContacto(request, id=None):
  if request.method == 'DELETE':
      if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
          return HttpResponseRedirect('/')
      else:
          contacto = Contacto.objects.get(id=id)
          f_anterior = str(contacto.foto)
          if f_anterior != "contactos/no_photo.png" :
             try:
                os.remove("media/"+f_anterior)
             except OSError as e:
                 print(e)

          contacto.delete()
          messages.success(request, 'Borrado')
          return HttpResponse(status=200)
  else:
      return HttpResponseRedirect("/")


########## TEMPLATE LISTADO DE TIPOS DE CONTACTO ###########

@login_required(login_url="login:login")
def tiposContactos(request):
    tipos = Tipo_contacto.objects.all()
    return render(request,'contactos/tipo_contacto_list.html',{'tipos': tipos })

########## TEMPLATE  CREAR TIPO DE CONTACTO ###########

@login_required(login_url="login:login")
def creartipoContacto(request):
    return render(request,'contactos/nuevo_tipo_contacto.html')

############## FUNCION ACTUALIZAR TIPO CONTACTO #######################

@login_required(login_url="login:login")
def guardartipoContacto(request):
    if request.method == 'POST':
        t = Tipo_contacto()
        t.agente_id = request.user.id
        t.tipo = request.POST['tipo'].encode('ascii', 'ignore')
        t.save()
        messages.success(request, 'Creado')
        return HttpResponseRedirect('/mantenimientos/tiposContactos')
    else:
        return HttpResponseRedirect('/')


########## TEMPLATE EDITAR TIPO DE CONTACTO ###########

@login_required(login_url="login:login")
def editartipoContacto(request,id = None):
    tipo_contacto = Tipo_contacto.objects.get(id=id)
    return render(request,'contactos/editar_tipo_contacto.html',{'tipo_contacto': tipo_contacto })


############## FUNCION ACTUALIZAR TIPO CONTACTO #######################

@login_required(login_url="login:login")
def actualizartipoContacto(request):
    if request.method == 'POST':
        id = request.POST['tipo_id']
        tipo = Tipo_contacto.objects.get(id = id)
        tipo.tipo = request.POST['tipo'].encode('ascii', 'ignore')
        tipo.save()
        messages.success(request, 'Actualizado')
        return HttpResponseRedirect('/mantenimientos/tiposContactos')
    else:
        return HttpResponseRedirect('/')



############## FUNCION ELIMINAR TIPO DE CONTACTO  #######################

@csrf_exempt
@transaction.atomic
@login_required(login_url="login:login")
def eliminartipoContacto(request, id=None):
  if request.method == 'DELETE':
      tipo = Tipo_contacto.objects.get(id = id)
      tipo.delete()
      messages.success(request, 'Borrado')
      return HttpResponse(status=200)
  else:
      return HttpResponseRedirect("/")


############## FUNCION EDITAR MULTIPLES CONTACTOS  #######################

@login_required(login_url="login:login")
def editarContactos(request):
    if request.method == 'POST':
        if (User.objects.filter(pk=request.user.id, groups__name='administradores').exists()):
            return HttpResponseRedirect('/')
        else:
            contactos = request.POST['lista_contactos_editar']
            contactos = contactos.split(",")
            palabras = request.POST['palabras_clave']

            for i in contactos:
                contacto = Contacto.objects.filter(id=i).update(palabras = palabras)
            messages.success(request, 'Actualizados')
            return HttpResponseRedirect('/contactos')
    else:
        return HttpResponseRedirect('/')

############## FUNCION ELIMINAR MULTIPLES CONTACTOS  #######################

@login_required(login_url="login:login")
def eliminarContactos(request):
    if request.method == 'POST':
        if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
            return HttpResponseRedirect('/')
        else:
            contactos = request.POST['lista_contactos']
            contactos = contactos.split(",")

            for i in contactos:
                contacto = Contacto.objects.filter(id=i).delete()

            messages.success(request, 'Borrados')
            return HttpResponseRedirect('/contactos')
    else:
        return HttpResponseRedirect('/')

############## FUNCION IMPORTAR CONTACTOS  #######################

import os
import openpyxl
@login_required(login_url="login:login")
def importarContactos(request):
    if (User.objects.filter(pk=request.user.id, groups__name='administradores').exists()):
        return HttpResponseRedirect('/')
    else:
        msge = "Importados"
        try:
            doc = openpyxl.load_workbook('media/agentes/'+str(request.user.id)+'/contactos.xlsx')

            try:
                hoja = doc.get_sheet_by_name('CONTACTOS')
            except:
                msge = "Archivo Incorrecto"
                messages.success(request, msge)
                return HttpResponseRedirect('/')


            ultimafilahoja = hoja.max_row
            ultimafila = 1
            for r in range (1, ultimafilahoja):
                if not hoja.cell(row = r, column = 8).value is None:
                    ultimafila = r + 1
                else: break
            ultima = ultimafila - 1
            for celda in range (ultima) :
                celda = celda + 1
                if celda >= 3:
                    contacto = Contacto()
                    n = Notas()
                    contacto.nombres =  hoja.cell(row=celda,column=1).value
                    contacto.apellidos =  hoja.cell(row=celda,column=2).value
                    contacto.genero = hoja.cell(row=celda,column=3).value
                    contacto.dni =  hoja.cell(row=celda,column=4).value
                    contacto.foto = "contactos/no_photo.png"
                    contacto.agen_contacto = int(request.user.id)
                    contacto.palabras =  hoja.cell(row=celda,column=5).value
                    contacto.correo =  hoja.cell(row=celda,column=6).value
                    contacto.celular =  hoja.cell(row=celda,column=7).value
                    contacto.t_oficina =  hoja.cell(row=celda,column=8).value
                    contacto.ext =  hoja.cell(row=celda,column=9).value
                    if contacto.genero == "F":
                        contacto.saludo = "Estimada"
                        contacto.carta = "Sra"
                    elif contacto.genero == "M":
                        contacto.saludo = "Estimado"
                        contacto.carta = "Sr"
                    contacto.save()
                    n.agente = request.user.id
                    n.usuario = contacto.id
                    mensaje = ("Creacion del Contacto")
                    n.nota = mensaje.encode('ascii', 'ignore')
                    n.save()
                    c = Contacto.objects.get(id=contacto.id)
                    c.nota_id = n.id
                    c.save()

                    try:
                        os.remove("media/agentes/"+str(request.user.id)+"/contactos.xlsx")
                    except OSError as e:
                        print(e)

        except IOError as e:
            msge = "No Encontrado"

        messages.success(request, msge)
        return HttpResponseRedirect('/contactos')


################### TEMPLATE CREAR NOTA  #######################

@login_required(login_url="login:login")
def crearNota(request,id = None):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        contacto = Contacto.objects.get(id = id)
        notas = Notas.objects.filter(usuario=id)
        url = URL
        return render(request,'contactos/nueva_nota.html',{'contacto': contacto,'notas': notas , 'url': url })


################### FUNCION CREAR NOTA  #######################

@csrf_exempt
@login_required(login_url="login:login")
def guardarNota(request):
    if request.method == 'PUT':
        if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
            return HttpResponseRedirect('/')
        else:
            n = Notas()
            n.usuario = int(request.GET['contacto_id'])
            n.agente = request.user.id
            nota = request.GET['txt_notas']
            n.nota = nota.encode('ascii', 'ignore')
            n.save()
            messages.success(request, 'Creada')
            return HttpResponse("Creada", status=200)
    else:
        return HttpResponseRedirect('/')


########## TEMPLATE SUBIR EXEL CONTACTOS ###########

@login_required(login_url="login:login")
def importarExcel(request):
    if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
        return HttpResponseRedirect('/')
    else:
        return render(request,'contactos/importar_excel.html')


############ FUNCION SUBIR EXEL DE CONTACTOS ###############

def upload_file(request):
    if request.method == 'POST':
        if User.objects.filter(pk=request.user.id, groups__name='administradores').exists():
            return HttpResponseRedirect('/')
        else:
            try:
                os.remove("media/agentes/"+str(request.user.id)+"/contactos.xlsx")
            except OSError as e:
                print(e)

            doc = Document()
            doc.filename = request.user.id
            doc.docfile = request.FILES['styled_file']
            doc.save()
            messages.success(request, 'Cargado')
            return HttpResponseRedirect('/contactos')
    else:
        return HttpResponseRedirect('/')
